#!/usr/bin/env python3
"""업무 관리 양식 세트 (엑셀)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

wb = Workbook()

NAVY = "1F3A5F"
LIGHT = "F0F4F8"
YELLOW = "FFF8E1"
GREEN_OK = "C8E6C9"
RED_BAD = "FFCDD2"
BORDER = Side(style="thin", color="999999")
FULL_BORDER = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)

H_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor=NAVY)
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

T_FONT = Font(name="맑은 고딕", size=10)
T_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color=NAVY)
SUB_FONT = Font(name="맑은 고딕", size=10, italic=True, color="666666")
LABEL_FILL = PatternFill("solid", fgColor=LIGHT)
LABEL_FONT = Font(name="맑은 고딕", size=10, bold=True)


def write_title(ws, title, subtitle, span=8):
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(2, 1, subtitle).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


def write_headers(ws, headers, row=4):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = H_ALIGN
        cell.border = FULL_BORDER
    ws.row_dimensions[row].height = 26


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER


# ───────────────── A. 업무 요청서 ─────────────────
ws = wb.active
ws.title = "A_업무요청서"
write_title(ws, "[양식 A] 업무 요청서", "신규 업무는 반드시 본 양식으로 접수합니다. 구두/메신저 요청도 수령자가 본 시트에 등록합니다.")
fields = [
    ("요청 번호", "REQ-YYYYMMDD-001"),
    ("요청일자", ""),
    ("요청자 (이름/소속)", ""),
    ("담당 희망자 (지명 가능 시)", ""),
    ("업무 제목", ""),
    ("업무 설명 (배경·목적)", ""),
    ("기대 산출물 (구체적 결과물)", ""),
    ("희망 마감일", ""),
    ("P등급 (P1~P4)", "P2"),
    ("P등급 사유", ""),
    ("예상 영향 (미수행 시 리스크)", ""),
    ("참고 자료/링크", ""),
    ("====== 이하 담당자 작성 ======", ""),
    ("배정 담당자", ""),
    ("예상 소요시간 (h)", ""),
    ("확정 마감일", ""),
    ("상태 (접수/진행/완료/반려)", "접수"),
    ("비고", ""),
]
for i, (label, default) in enumerate(fields, start=4):
    a = ws.cell(i, 1, label)
    b = ws.cell(i, 2, default)
    a.font = LABEL_FONT
    a.fill = LABEL_FILL
    a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    a.border = FULL_BORDER
    b.font = T_FONT
    b.alignment = T_ALIGN
    b.border = FULL_BORDER
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    if "====" in label:
        a.fill = PatternFill("solid", fgColor=YELLOW)
set_widths(ws, [28, 30, 30, 30])
dv = DataValidation(type="list", formula1='"P1,P2,P3,P4"', allow_blank=True)
dv.add(ws.cell(12, 2))
ws.add_data_validation(dv)
dv2 = DataValidation(type="list", formula1='"접수,진행,완료,반려,보류"', allow_blank=True)
dv2.add(ws.cell(20, 2))
ws.add_data_validation(dv2)

# ───────────────── B. 통합 업무 보드 ─────────────────
ws = wb.create_sheet("B_통합업무보드")
write_title(
    ws,
    "[양식 B] 통합 업무 보드",
    "전사 모든 업무의 단일 저장소. 보드에 등록되지 않은 업무는 \"존재하지 않는 업무\"로 간주합니다.",
    span=12,
)
headers = [
    "ID", "담당자", "업무명", "유형", "P등급",
    "예상 소요(h)", "실제 소요(h)", "시작일", "마감일", "상태", "요청자", "비고",
]
write_headers(ws, headers, row=4)
sample = [
    ["BD-0001", "홍길동", "5월 정기 매출보고서 작성", "정기", "P2", 3, "", "2026-05-11", "2026-05-13", "진행", "경영지원팀장", ""],
    ["BD-0002", "김철수", "신상품 A 상세페이지 디자인", "비정기", "P2", 6, "", "2026-05-11", "2026-05-15", "진행", "MD", ""],
    ["BD-0003", "이영희", "5월 광고소재 제작", "정기", "P3", 4, "", "2026-05-11", "2026-05-17", "접수", "마케팅팀장", ""],
]
for r, row in enumerate(sample, start=5):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER
for r in range(8, 60):
    for c in range(1, 13):
        ws.cell(r, c).border = FULL_BORDER
        ws.cell(r, c).font = T_FONT
        ws.cell(r, c).alignment = T_ALIGN
set_widths(ws, [10, 10, 30, 8, 7, 9, 9, 12, 12, 10, 12, 20])
# 상태 dropdown + P등급 dropdown + 유형
dv_status = DataValidation(type="list", formula1='"접수,진행,완료,반려,보류"', allow_blank=True)
dv_status.add(f"J5:J60")
ws.add_data_validation(dv_status)
dv_p = DataValidation(type="list", formula1='"P1,P2,P3,P4"', allow_blank=True)
dv_p.add(f"E5:E60")
ws.add_data_validation(dv_p)
dv_type = DataValidation(type="list", formula1='"정기,비정기,개선,학습"', allow_blank=True)
dv_type.add(f"D5:D60")
ws.add_data_validation(dv_type)

# 마감 임박/지연 조건부 서식
red_fill = PatternFill("solid", fgColor=RED_BAD)
yellow_fill = PatternFill("solid", fgColor=YELLOW)
ws.conditional_formatting.add(
    "I5:I60",
    FormulaRule(formula=[f'AND(I5<>"", I5<TODAY(), J5<>"완료")'], fill=red_fill),
)
ws.conditional_formatting.add(
    "I5:I60",
    FormulaRule(formula=[f'AND(I5<>"", I5-TODAY()<=2, I5>=TODAY(), J5<>"완료")'], fill=yellow_fill),
)
ws.freeze_panes = "A5"

# ───────────────── C. 데일리 스탠드업 회의록 ─────────────────
ws = wb.create_sheet("C_데일리스탠드업")
write_title(ws, "[양식 C] 데일리 스탠드업 회의록", "매일 09:00–09:15 / 1인 1~2분 / 토론 금지", span=6)
ws.cell(4, 1, "회의 일자").font = LABEL_FONT
ws.cell(4, 1).fill = LABEL_FILL
ws.cell(4, 2, "")
ws.cell(4, 3, "진행자").font = LABEL_FONT
ws.cell(4, 3).fill = LABEL_FILL
ws.cell(4, 4, "")
ws.cell(4, 5, "기록자").font = LABEL_FONT
ws.cell(4, 5).fill = LABEL_FILL
ws.cell(4, 6, "")
for c in range(1, 7):
    ws.cell(4, c).border = FULL_BORDER
write_headers(ws, ["참석자", "어제 한 일", "오늘 할 일", "막힌 점·도움 요청", "관련 BD-ID", "후속 조치자"], row=6)
for r in range(7, 20):
    style_data_row(ws, r, 6)
set_widths(ws, [12, 35, 35, 30, 12, 12])

# 막힌 이슈 후속 액션
ws.cell(22, 1, "막힌 이슈 후속 액션 (스탠드업 후 별도 30분 내 처리)").font = LABEL_FONT
ws.cell(22, 1).fill = PatternFill("solid", fgColor=YELLOW)
ws.merge_cells(start_row=22, start_column=1, end_row=22, end_column=6)
write_headers(ws, ["이슈", "해결책 합의", "담당자", "기한", "상태", "비고"], row=23)
for r in range(24, 30):
    style_data_row(ws, r, 6)

# ───────────────── D. 표준 업무 공수표 ─────────────────
ws = wb.create_sheet("D_표준공수표")
write_title(ws, "[양식 D] 표준 업무 공수표", "업무 유형별 1회 처리 표준 시간. 누적 시간 산정·80% 룰의 근거 데이터.", span=6)
write_headers(ws, ["업무코드", "부문", "업무명", "단위", "표준 소요(h)", "비고"], row=4)
samples = [
    ["MK-001", "마케팅", "광고 소재 1세트 제작", "1세트", 4, "이미지 3종 + 카피"],
    ["MK-002", "마케팅", "주간 광고 리포트", "1건", 2, "주 1회 정기"],
    ["MK-003", "마케팅", "콘텐츠 기획 + 발행 1건", "1건", 6, "SNS 기준"],
    ["DV-001", "개발", "신규 페이지 퍼블리싱", "1page", 8, "디자인 시안 받은 후"],
    ["DV-002", "개발", "버그 수정 (Minor)", "1건", 1, ""],
    ["DV-003", "개발", "버그 수정 (Major)", "1건", 4, ""],
    ["MD-001", "MD", "신상품 1개 등록", "1건", 2, "사진·설명·옵션 포함"],
    ["MD-002", "MD", "주간 매입 발주", "1주", 3, ""],
    ["VM-001", "VMD", "오프라인 매장 진열 변경", "1매장", 6, ""],
    ["DS-001", "디자인", "상세페이지 디자인", "1건", 8, ""],
    ["DS-002", "디자인", "배너 디자인", "1건", 2, ""],
    ["CS-001", "CS", "1:1 문의 응대", "1건", 0.25, "평균"],
    ["CS-002", "CS", "반품/교환 처리", "1건", 0.5, ""],
    ["SL-001", "영업", "B2B 신규 미팅", "1건", 3, "이동시간 포함"],
    ["SL-002", "영업", "거래처 정기 컨택", "1건", 0.5, ""],
    ["HR-001", "인사", "급여 마감", "1회", 8, "월 1회"],
    ["HR-002", "인사", "채용 1포지션 진행", "1포지션", 20, "공고~입사까지 누적"],
    ["AC-001", "회계", "월 결산", "1회", 16, "월 1회"],
    ["GA-001", "총무", "법인카드 정산", "1회", 4, "월 1회"],
]
for r, row in enumerate(samples, start=5):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER
set_widths(ws, [10, 10, 35, 12, 12, 30])

# 가용시간 산정 영역
start_row = 5 + len(samples) + 3
ws.cell(start_row, 1, "■ 개인별 주간 가용시간 산정 (80% 룰)").font = TITLE_FONT
ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
write_headers(ws, ["담당자", "주 법정시간(h)", "버퍼(20%)", "가용 상한(h)", "이번주 누적 예상(h)", "활용률"], row=start_row + 1)
people = ["마케팅팀장", "컨텐츠마케터", "퍼포먼스마케터", "프론트엔드", "MD", "VMD", "디자이너",
          "CS", "인사", "회계", "총무", "영업", "경영지원팀장"]
for i, name in enumerate(people):
    r = start_row + 2 + i
    ws.cell(r, 1, name).font = T_FONT
    ws.cell(r, 2, 40)
    ws.cell(r, 3, f"=B{r}*0.2")
    ws.cell(r, 4, f"=B{r}-C{r}")
    ws.cell(r, 5, 0)  # 사용자가 입력
    ws.cell(r, 6, f"=IF(D{r}=0,0,E{r}/D{r})")
    ws.cell(r, 6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(r, c).font = T_FONT
        ws.cell(r, c).alignment = T_ALIGN
        ws.cell(r, c).border = FULL_BORDER

# 활용률 조건부 서식
util_range = f"F{start_row + 2}:F{start_row + 1 + len(people)}"
ws.conditional_formatting.add(util_range, CellIsRule(operator="greaterThan", formula=["1"], fill=red_fill))
ws.conditional_formatting.add(util_range, CellIsRule(operator="between", formula=["0.85", "1"], fill=yellow_fill))
ws.conditional_formatting.add(util_range, CellIsRule(operator="lessThanOrEqual", formula=["0.85"], fill=PatternFill("solid", fgColor=GREEN_OK)))

# ───────────────── E. 역할 로테이션 표 ─────────────────
ws = wb.create_sheet("E_역할로테이션")
write_title(ws, "[양식 E] 역할 로테이션 표", "팀장 부재 시 팀장 기능을 4개 역할로 분산하여 순환합니다.", span=6)
write_headers(ws, ["주차", "기간", "업무조정자", "진행자", "기록자", "외부 창구"], row=4)
weeks = [
    ["W19", "05/05–05/11", "마케팅팀장", "컨텐츠마케터", "퍼포먼스마케터", "경영지원팀장"],
    ["W20", "05/12–05/18", "마케팅팀장", "퍼포먼스마케터", "디자이너", "경영지원팀장"],
    ["W21", "05/19–05/25", "마케팅팀장", "MD", "VMD", "경영지원팀장"],
    ["W22", "05/26–06/01", "경영지원팀장", "프론트엔드", "CS", "경영지원팀장"],
    ["W23", "06/02–06/08", "경영지원팀장", "인사", "회계", "마케팅팀장"],
    ["W24", "06/09–06/15", "경영지원팀장", "총무", "영업", "마케팅팀장"],
]
for r, row in enumerate(weeks, start=5):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER
set_widths(ws, [8, 16, 16, 16, 16, 16])

ws.cell(13, 1, "역할 정의").font = TITLE_FONT
write_headers(ws, ["역할", "책임", "임기"], row=14)
defs = [
    ["업무조정자", "신규 업무 접수·배정, 보드 정합성 점검, 80% 룰 모니터링", "2주~1개월"],
    ["진행자", "데일리/위클리 회의 진행, 시간 관리", "1주"],
    ["기록자", "회의록 작성, 결정사항 보드 반영", "1주"],
    ["외부 창구", "타 부서·상위 보고 단일 채널, 외부 요청 1차 수신", "1개월"],
]
for r, row in enumerate(defs, start=15):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER
ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=5)
ws.merge_cells(start_row=14, start_column=6, end_row=14, end_column=6)

# ───────────────── F. 주간 회고록 ─────────────────
ws = wb.create_sheet("F_주간회고록")
write_title(ws, "[양식 F] 주간 회고록", "매주 금요일 30분. 사람 탓 아닌 시스템 개선 안건 도출.", span=5)
header_rows = [
    ("회고 주차", ""),
    ("진행자", ""),
    ("기록자", ""),
    ("참석자", ""),
]
for i, (a, b) in enumerate(header_rows, start=4):
    ws.cell(i, 1, a).font = LABEL_FONT
    ws.cell(i, 1).fill = LABEL_FILL
    ws.cell(i, 1).border = FULL_BORDER
    ws.cell(i, 2, b).border = FULL_BORDER
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

ws.cell(9, 1, "■ 1. 실적 리뷰").font = TITLE_FONT
write_headers(ws, ["BD-ID", "업무명", "담당자", "완료 여부", "미완료 사유 분류"], row=10)
for r in range(11, 20):
    style_data_row(ws, r, 5)
    dv = DataValidation(type="list", formula1='"완료,부분완료,미완료,이월"', allow_blank=True)
    dv.add(ws.cell(r, 4))
    ws.add_data_validation(dv)
    dv2 = DataValidation(type="list", formula1='"양 문제,역량 문제,프로세스 문제,외부 의존,N/A"', allow_blank=True)
    dv2.add(ws.cell(r, 5))
    ws.add_data_validation(dv2)

ws.cell(21, 1, "■ 2. 좋았던 점 (Keep)").font = TITLE_FONT
for r in range(22, 25):
    ws.cell(r, 1, "").border = FULL_BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

ws.cell(26, 1, "■ 3. 문제점 (Problem)").font = TITLE_FONT
for r in range(27, 30):
    ws.cell(r, 1, "").border = FULL_BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

ws.cell(31, 1, "■ 4. 개선 액션 (Try) — 다음 주 적용").font = TITLE_FONT
write_headers(ws, ["No.", "개선 액션", "담당자", "기한", "성공 기준"], row=32)
for r in range(33, 38):
    style_data_row(ws, r, 5)

set_widths(ws, [12, 35, 14, 14, 25])

# ───────────────── G. 연장근로 사전동의서 ─────────────────
ws = wb.create_sheet("G_연장근로동의서")
write_title(ws, "[양식 G] 연장근로 사전동의서", "근로기준법 §53, §56에 따라 사전 동의 + 1.5배 가산수당 지급.", span=4)
fields = [
    ("연도/월/주차", ""),
    ("성명", ""),
    ("소속", ""),
    ("주 법정근로시간(h)", "40"),
    ("이번 주 기 등록 업무 누적(h)", ""),
    ("추가 연장근로 사유 (구체적 업무)", ""),
    ("예상 연장시간(h)", ""),
    ("연장 일정 (일자/시간대)", ""),
    ("주간 총 근로시간 합계", "=B7+B10"),
    ("52시간 초과 여부 (자동)", '=IF(B12>52,"초과 - 승인 불가","적정")'),
    ("연장근로수당 산정(통상시급 × 시간 × 1.5)", ""),
    ("본인 동의 (서명/날인)", ""),
    ("부서장 승인", ""),
    ("인사팀 확인", ""),
    ("비고", ""),
]
for i, (a, b) in enumerate(fields, start=4):
    ws.cell(i, 1, a).font = LABEL_FONT
    ws.cell(i, 1).fill = LABEL_FILL
    ws.cell(i, 1).border = FULL_BORDER
    cell = ws.cell(i, 2, b)
    cell.border = FULL_BORDER
    cell.font = T_FONT
    cell.alignment = T_ALIGN
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

# 52시간 초과 자동 표시 조건부 서식
ws.conditional_formatting.add(
    "B13", FormulaRule(formula=['B13="초과 - 승인 불가"'], fill=red_fill, font=Font(name="맑은 고딕", bold=True, color="B71C1C"))
)
set_widths(ws, [32, 20, 20, 20])

# ───────────────── 안내 ─────────────────
ws = wb.create_sheet("README", 0)
write_title(ws, "업무 관리 양식 세트", "팀장 부재·주니어 중심 조직의 운영체계", span=4)
ws.cell(4, 1, "본 파일은 「인사팀장_미팅자료_업무프로세스개선안.pdf」의 별첨입니다.").font = Font(name="맑은 고딕", size=11)
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
write_headers(ws, ["코드", "양식명", "사용 시점", "작성 책임자"], row=6)
forms = [
    ["A", "업무 요청서", "신규 업무 발생 시", "요청자"],
    ["B", "통합 업무 보드", "상시 (모든 업무)", "담당자"],
    ["C", "데일리 스탠드업 회의록", "매일 09:00", "기록자(주 로테이션)"],
    ["D", "표준 공수표 + 가용시간 산정", "월 1회 갱신 + 매주 누적 입력", "팀 공동"],
    ["E", "역할 로테이션 표", "분기 단위 사전 공시", "업무조정자"],
    ["F", "주간 회고록", "매주 금요일", "기록자"],
    ["G", "연장근로 사전동의서", "80% 룰 초과 시", "담당자 + 인사"],
]
for r, row in enumerate(forms, start=7):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = T_FONT
        cell.alignment = T_ALIGN
        cell.border = FULL_BORDER
set_widths(ws, [8, 30, 30, 25])

ws.cell(16, 1, "■ 운영 원칙").font = TITLE_FONT
principles = [
    "1. 보드에 등록되지 않은 업무는 \"존재하지 않는 업무\"로 간주합니다.",
    "2. 가용 시간의 80%(주 32h)까지만 업무를 채웁니다. 나머지는 돌발·휴게·이동 버퍼.",
    "3. 모든 회의는 사전 의제 + 시간 박싱 + 결정 1건 이상의 산출물을 남깁니다.",
    "4. 퇴근 후 메신저로 업무 지시를 하지 않습니다. 보드 마감일은 근무시간 내로 설정합니다.",
    "5. PC 감시·키로깅 등의 도구는 사용하지 않습니다 (개인정보보호법 위반 소지).",
]
for i, t in enumerate(principles, start=17):
    ws.cell(i, 1, t).font = T_FONT
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

# Save
wb.save("/home/user/Jay/업무관리_양식세트.xlsx")
print("XLSX generated successfully")
